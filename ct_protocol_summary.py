import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, simpledialog

# Function to open file dialog for HTML file selection
def select_html_file():
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(title="Select the HTML File", filetypes=[("HTML files", "*.html")])
    return file_path

# Function to open file dialog for Excel file save location
def select_save_location():
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    save_path = filedialog.askdirectory(title="Select Save Location for Excel File")
    return save_path

# Get the HTML file path, save location, and file name
html_file_path = select_html_file()
save_location = select_save_location()
file_name = simpledialog.askstring("Input", "Enter the Excel file name (without extension):")

# Ensure valid file paths and file name
if not html_file_path or not save_location or not file_name:
    print("Invalid inputs provided. Exiting...")
    exit()

# Full output file path
output_file = f"{save_location}/{file_name}.xlsx"

# Read the HTML file
with open(html_file_path, 'r', encoding='utf-8') as file:
    html_content = file.read()

# Parse the HTML using BeautifulSoup
soup = BeautifulSoup(html_content, 'html.parser')

# Initialize lists to hold the combined data
combined_data = []
headers = set()

# Loop through the acquisition labels and tables
current_acquisition = None
current_protocol = None
acquisition_number = None
in_result_section = False

for element in soup.find_all(['p', 'table']):
    if element.name == 'p' and 'Acquisition label' in element.get_text():
        # Start a new acquisition record
        current_protocol = element.find_previous('p', class_='exam').get_text(strip=True)
        acquisition_number = element.get_text(strip=True).split(',')[0].replace('Acquisition label :', '').strip()
        label = element.get_text(strip=True).split(',')[1].strip() if ',' in element.get_text(strip=True) else ''
        current_acquisition = {
            'Protocol': current_protocol,
            'Acquisition Number': acquisition_number,
            'Label': label,
            'Type': 'Acquisition',
            'Result Label': ''
        }
        combined_data.append(current_acquisition)
        in_result_section = False  # Reset the flag since we're in a new acquisition section

    elif element.name == 'table' and current_acquisition is not None and not in_result_section:
        # Extract parameters and values from the acquisition table
        rows = element.find_all('tr')
        for row in rows:
            cells = row.find_all('td')
            if len(cells) == 2:
                parameter = cells[0].get_text(strip=True)
                value = cells[1].get_text(strip=True)
                current_acquisition[parameter] = value
                headers.add(parameter)

    elif element.name == 'p' and 'Result Label' in element.get_text():
        # Start a new result record under the current acquisition
        result_label = element.get_text(strip=True).replace('Result Label :', '').strip()
        current_result = {
            'Protocol': current_protocol,
            'Acquisition Number': acquisition_number,
            'Label': '',
            'Type': 'Result',
            'Result Label': result_label
        }
        combined_data.append(current_result)
        in_result_section = True  # We are now entering a result section under the acquisition

    elif element.name == 'table' and in_result_section:
        # Extract parameters and values from the result table
        rows = element.find_all('tr')
        for row in rows:
            cells = row.find_all('td')
            if len(cells) == 2:
                parameter = cells[0].get_text(strip=True)
                value = cells[1].get_text(strip=True)
                current_result[parameter] = value
                headers.add(parameter)

# Convert the combined data into a DataFrame
df = pd.DataFrame(combined_data)

# Reorder the columns: Put the essential ones (Protocol, Acquisition Number, etc.) at the front
essential_columns = ['Protocol', 'Acquisition Number', 'Label', 'Type', 'Result Label']
remaining_columns = [col for col in df.columns if col not in essential_columns]

# Reindex the DataFrame with essential columns first and remaining columns afterward
df = df[essential_columns + remaining_columns]

# Convert each column to numeric if possible, catching exceptions for non-numeric columns
for col in remaining_columns:
    try:
        df[col] = pd.to_numeric(df[col])
    except ValueError:
        # If a column contains non-numeric data, it will be left unchanged
        pass

# Save the DataFrame to an Excel file
df.to_excel(output_file, index=False)

# Now apply formatting using openpyxl
wb = load_workbook(output_file)
ws = wb.active

# Increase header height
ws.row_dimensions[1].height = 30  # Adjust the height of the header row

# Auto-fit the width of all columns
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column letter
    for cell in col:
        try:  # Necessary to avoid errors for empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)  # Add a bit of extra space
    ws.column_dimensions[column].width = adjusted_width

# Save the workbook with the formatting applied
wb.save(output_file)

print(f"Data successfully extracted and saved to {output_file} with formatting")